#!/usr/bin/env python3
"""
Accessibility issue tracker.

Builds a spreadsheet of accessibility-labeled GitHub issues:
  * every open one across all public repos, ranked by repo popularity
  * every one in your own repos, open and closed
  * a per-repo rollup of your own repos

Auth is optional. GitHub's search API answers unauthenticated requests, so a
plain `python3 track_a11y.py --owner someone` works with no sign-in at all.
If a token is available (GH_TOKEN, or a signed-in `gh`), the tool uses it for
higher rate limits, private repos, and the far cheaper GraphQL path.

Usage:
    python3 track_a11y.py --owner your-name
    python3 track_a11y.py --owner your-name --global-limit 500
    python3 track_a11y.py --owner your-name --no-global      # skip the public scan
    python3 track_a11y.py --owner your-name,your-org
"""

import argparse
import json
import os
import re
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone

API = "https://api.github.com"

# Label spellings that all mean "accessibility". GitHub treats a comma-separated
# label qualifier as OR, so these go into one query rather than one query each.
DEFAULT_LABELS = ["accessibility", "a11y", "accessibility-issue", "accessibility issue"]

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUT = os.path.join(SCRIPT_DIR, "accessibility-issues.xlsx")

# Without a token the core API allows 60 requests an hour, so the number of
# repos we look up stars for has to stay small.
ANON_STAR_LOOKUPS = 25

# Owners skipped by default on the public sheet. These are companies with paid
# accessibility teams and the budget to fix their own issues; volunteer effort is
# better spent on independent projects. Edit this list freely, or pass
# --include-big-tech to switch the filter off entirely.
BIG_TECH_OWNERS = {
    # Microsoft
    "microsoft", "github", "azure", "dotnet", "azure-samples", "microsoftdocs",
    "typescript", "playwright-community", "visualstudio", "cli",
    # Google
    "google", "googleapis", "googlechrome", "googlecloudplatform", "google-research",
    "angular", "tensorflow", "flutter", "firebase", "chromium", "grpc", "dart-lang",
    "golang", "bazelbuild", "material-components",
    # Meta
    "facebook", "facebookresearch", "facebookincubator", "meta-llama", "reactjs",
    "pytorch", "react-native-community", "react", "reactwg",
    # Apple, Amazon
    "apple", "swiftlang", "aws", "awslabs", "amzn", "aws-samples", "amazon-archives",
    # Other large tech
    "netflix", "adobe", "oracle", "mysql", "ibm", "redhat", "openshift", "salesforce",
    "sap", "intel", "nvidia", "uber", "uber-go", "airbnb", "shopify", "stripe",
    "twitter", "x", "linkedin", "atlassian", "cloudflare", "elastic", "mongodb",
    "docker", "jetbrains", "kotlin", "spotify", "bytedance", "alibaba", "tencent",
    "baidu", "huawei", "samsung", "sony", "dell", "cisco", "vmware", "canonical",
    "palantir", "snapchat", "paypal", "block", "square", "zoom", "slackapi",
    "dropbox", "reddit", "discord", "epicgames", "unity-technologies", "roblox",
    "twilio", "zendesk", "hashicorp", "datadog", "newrelic", "splunk", "grafana",
    "confluentinc", "snowflakedb", "databricks", "openai", "anthropics", "vercel",
    "supabase", "atlassian-labs", "expo", "sentry", "gitlab-org",
}

# GitHub reports a license for almost everything with a LICENSE file. NOASSERTION
# means it found one it could not identify, which is not a promise of open source.
UNCLEAR_LICENSES = {None, "", "NOASSERTION"}

# Accessibility specialties, checked in order: the first match wins, so the more
# specific areas have to come before the general ones. Someone who does screen
# reader work and someone who does color contrast work want different rows, and
# neither wants to read all 300.
AREAS = [
    ("Screen reader & ARIA", [
        "screen reader", "screenreader", "voiceover", "nvda", "jaws", "talkback",
        "narrator", "aria-", "aria ", "role=", "landmark", "semantics", "announce",
        "alt text", "alt attribute", "image description", "accessible name",
    ]),
    ("Keyboard & focus", [
        "keyboard", "focus", "tab order", "tabindex", "tab key", "shortcut",
        "focus trap", "focus ring", "focus indicator", "arrow key",
    ]),
    ("Captions & media", [
        "caption", "subtitle", "transcript", "audio description", "sign language",
    ]),
    ("Motion & seizure safety", [
        "reduced motion", "prefers-reduced-motion", "animation", "parallax",
        "flashing", "blinking", "seizure", "autoplay",
    ]),
    ("Color & contrast", [
        "contrast", "color blind", "colour blind", "colorblind", "low vision",
        "dark mode", "light mode", "high contrast", "color only", "colour only",
    ]),
    ("Text & zoom", [
        "font size", "text size", "zoom", "text scaling", "line height",
        "dyslexi", "reflow", "magnif",
    ]),
    ("Forms & error messages", [
        "form field", "form control", "form label", "input field", "placeholder",
        "error message", "validation", "required field", "checkbox", "radio button",
        "autocomplete attribute", "label for",
    ]),
    ("Touch & mobile", [
        "touch target", "tap target", "gesture", "pinch", "swipe", "mobile a11y",
        "hit area",
    ]),
    ("Cognitive & plain language", [
        "cognitive", "plain language", "readability", "timeout", "time limit",
        "distraction", "easy read",
    ]),
]

# Phrases used to find accessibility work that nobody labeled. GitHub search
# treats a quoted phrase as an exact match, so these stay narrow on purpose:
# "focus" alone would return half of GitHub.
UNTAGGED_PHRASES = [
    "screen reader", "keyboard navigation", "color contrast",
    "alt text", "WCAG", "aria-label",
]


GENERAL = "General / unclassified"


# Matching is anchored to word starts. Substring matching quietly wrecked this:
# "form" fired on "information", "platform", and "performance", which put 87
# unrelated issues into Forms & error messages.
AREA_PATTERNS = [
    (area, re.compile(r"\b(?:" + "|".join(re.escape(k) for k in needles) + r")", re.I))
    for area, needles in AREAS
]


def classify_all(title, labels, body=""):
    """Every specialty an issue touches, in AREAS order.

    Issues are rarely about one thing: a modal dialog bug is usually both a
    focus problem and a screen reader problem. The first match becomes the
    primary area for grouping, the rest go in "Also covers" so filtering on any
    one of them finds the issue.
    """
    haystack = " ".join([title or "", " ".join(labels or []), (body or "")[:1500]])
    hits = [area for area, pattern in AREA_PATTERNS if pattern.search(haystack)]
    return hits or [GENERAL]


def classify(title, labels, body=""):
    return classify_all(title, labels, body)[0]


# --------------------------------------------------------------------------
# Auth: entirely optional
# --------------------------------------------------------------------------

def resolve_token():
    """A token if one is lying around, otherwise None. Never prompts."""
    for var in ("GH_TOKEN", "GITHUB_TOKEN"):
        if os.environ.get(var):
            return os.environ[var]
    try:
        proc = subprocess.run(["gh", "auth", "token"], capture_output=True, text=True, timeout=10)
        if proc.returncode == 0 and proc.stdout.strip():
            return proc.stdout.strip()
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    return None


TOKEN = None  # set in main()


# --------------------------------------------------------------------------
# HTTP
# --------------------------------------------------------------------------

RATE_LIMITED = False  # set once anonymous budget runs out, to stop retrying


def request(url, data=None, retries=3, soft=False):
    """soft=True means a rate limit returns None instead of ending the run.

    Used for the optional star lookups: losing a popularity number is worth far
    less than losing the whole spreadsheet.
    """
    global RATE_LIMITED
    if soft and RATE_LIMITED:
        return None
    headers = {
        "Accept": "application/vnd.github+json",
        "User-Agent": "a11y-tracker",
    }
    if TOKEN:
        headers["Authorization"] = f"Bearer {TOKEN}"
    body = json.dumps(data).encode() if data is not None else None
    req = urllib.request.Request(url, data=body, headers=headers)

    for attempt in range(retries):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                return json.loads(resp.read().decode())
        except urllib.error.HTTPError as err:
            if err.code in (403, 429) and soft:
                RATE_LIMITED = True
                return None
            # Search is limited to 10 requests a minute without a token, and
            # GitHub answers a burst with 403 rather than a 429.
            if err.code in (403, 429) and attempt < retries - 1:
                wait = int(err.headers.get("Retry-After") or (20 * (attempt + 1)))
                print(f"  rate limited, waiting {wait}s...", file=sys.stderr)
                time.sleep(wait)
                continue
            if err.code == 404:
                return None
            detail = err.read().decode(errors="replace")[:200]
            if err.code in (403, 429):
                hint = ("\nHint: set GH_TOKEN or run 'gh auth login' for a much higher limit."
                        if not TOKEN else "")
                sys.exit(f"GitHub rate limit hit.{hint}\n{detail}")
            sys.exit(f"GitHub API error {err.code} for {url}\n{detail}")
        except urllib.error.URLError as err:
            if attempt < retries - 1:
                time.sleep(3)
                continue
            sys.exit(f"Could not reach github.com: {err.reason}")
    return None


def graphql(query, variables):
    payload = request(f"{API}/graphql", data={"query": query, "variables": variables})
    if payload and payload.get("errors"):
        sys.exit("GitHub GraphQL error: " + json.dumps(payload["errors"])[:300])
    return (payload or {}).get("data")


def label_query(labels):
    """Quote multi-word labels, join with commas so GitHub reads it as OR."""
    return "label:" + ",".join(f'"{l}"' if " " in l else l for l in labels)


# --------------------------------------------------------------------------
# Searches
# --------------------------------------------------------------------------

def search_rest(query, limit, sort=None):
    """Paged REST issue search. GitHub caps any single search at 1000 results."""
    items, page = [], 1
    while len(items) < min(limit, 1000):
        params = {"q": query, "per_page": 100, "page": page}
        if sort:
            params["sort"], params["order"] = sort, "desc"
        data = request(f"{API}/search/issues?" + urllib.parse.urlencode(params))
        batch = (data or {}).get("items", [])
        items.extend(batch)
        if len(batch) < 100 or page >= 10:
            break
        page += 1
    return items[:limit]


GRAPHQL_SEARCH = """
query($q: String!, $after: String) {
  search(query: $q, type: ISSUE, first: 100, after: $after) {
    pageInfo { hasNextPage endCursor }
    nodes {
      ... on Issue {
        # No bodyText here on purpose: issue bodies are unbounded, and asking for
        # 100 of them either trips GraphQL's cost limit or times out with a 502.
        # Specialty is guessed from the title and labels instead, which is what
        # the labeled path has to work with anyway.
        number title url createdAt updatedAt state
        comments { totalCount }
        labels(first: 20) { nodes { name } }
        repository {
          nameWithOwner stargazerCount isArchived isFork
          licenseInfo { spdxId name }
          primaryLanguage { name }
        }
      }
    }
  }
}
"""


def search_global(labels, limit, big_owners, drop_big, want_license, min_stars, per_repo):
    """Open accessibility issues across all public repos, with repo star counts.

    With a token this is GraphQL, which returns stars, license, and archived
    status inline: 100 issues per call. Without one it falls back to REST plus a
    capped set of star lookups, because the anonymous core limit is 60 an hour.

    Big-company rows are tagged rather than dropped, so the spreadsheet can hide
    them behind a filter the reader can switch off. `limit` counts independent
    rows only, so tagging them does not eat the budget.
    """
    base = f"{label_query(labels)} is:issue is:open is:public"
    dropped = {"big tech": 0, "no clear license": 0, "archived or fork": 0,
               "below star floor": 0, "extra issues from a repo already listed": 0}
    per_repo_count = {}

    def is_big(repo_full):
        return repo_full.split("/")[0].lower() in big_owners

    def keep(repo_full, license_id, archived, fork, stars):
        if drop_big and is_big(repo_full):
            dropped["big tech"] += 1
            return False
        if archived or fork:
            dropped["archived or fork"] += 1
            return False
        if want_license and license_id is not None and license_id in UNCLEAR_LICENSES:
            dropped["no clear license"] += 1
            return False
        if stars is not None and stars < min_stars:
            dropped["below star floor"] += 1
            return False
        # One busy repo can otherwise fill the whole sheet, which hides the point
        # of the exercise: finding projects to help, not one project's backlog.
        if per_repo and per_repo_count.get(repo_full, 0) >= per_repo:
            dropped["extra issues from a repo already listed"] += 1
            return False
        per_repo_count[repo_full] = per_repo_count.get(repo_full, 0) + 1
        return True

    if TOKEN:
        rows, cursor, independent = [], None, 0
        # Most-commented first: a sample out of ~17k should favour issues people
        # actually care about rather than whatever sorts first by default.
        while independent < limit:
            data = graphql(GRAPHQL_SEARCH, {"q": base + " sort:comments-desc", "after": cursor})
            if not data:
                break
            block = data["search"]
            for n in block["nodes"]:
                if not n or not n.get("repository"):
                    continue
                repo = n["repository"]
                license_id = (repo.get("licenseInfo") or {}).get("spdxId")
                if not keep(repo["nameWithOwner"], license_id,
                            repo.get("isArchived"), repo.get("isFork"),
                            repo["stargazerCount"]):
                    continue
                big = is_big(repo["nameWithOwner"])
                independent += 0 if big else 1
                node_labels = [l["name"] for l in n["labels"]["nodes"]]
                rows.append({
                    "repo": repo["nameWithOwner"],
                    "big": big,
                    "area": classify(n["title"], node_labels),
                    "source": "labeled",
                    "stars": repo["stargazerCount"],
                    "license": (repo.get("licenseInfo") or {}).get("spdxId") or "none",
                    "language": (repo.get("primaryLanguage") or {}).get("name") or "-",
                    "number": n["number"],
                    "title": n["title"],
                    "url": n["url"],
                    "labels": [l["name"] for l in n["labels"]["nodes"]],
                    "comments": n["comments"]["totalCount"],
                    "createdAt": n["createdAt"],
                    "updatedAt": n["updatedAt"],
                })
            if not block["pageInfo"]["hasNextPage"]:
                break
            cursor = block["pageInfo"]["endCursor"]
        capped = False
    else:
        # REST search carries no license or archived flag, so anonymous runs can
        # only apply the owner filter. Pull extra to cover what gets dropped.
        raw = search_rest(base, min(limit * 2, 1000), sort="comments")
        rows = []
        for i in raw:
            full = "/".join(i["repository_url"].split("/")[-2:])
            if not keep(full, None, False, False, None):
                continue
            if len([r for r in rows if not r["big"]]) >= limit:
                break
            rows.append({
                "repo": full,
                "big": is_big(full),
                "area": classify(i["title"], [l["name"] for l in i.get("labels", [])],
                                 i.get("body")),
                "source": "labeled",
                "stars": None,
                "license": "?",
                "language": "-",
                "number": i["number"],
                "title": i["title"],
                "url": i["html_url"],
                "labels": [l["name"] for l in i.get("labels", [])],
                "comments": i.get("comments", 0),
                "createdAt": i["created_at"],
                "updatedAt": i["updated_at"],
            })

        # Star counts cost one call each, so only the repos with the most hits
        # get looked up. The rest sort to the bottom with a blank star cell.
        by_hits = {}
        for r in rows:
            by_hits[r["repo"]] = by_hits.get(r["repo"], 0) + 1
        top = sorted(by_hits, key=lambda k: -by_hits[k])[:ANON_STAR_LOOKUPS]
        stars = {}
        with ThreadPoolExecutor(max_workers=4) as pool:
            lookup = lambda r: request(f"{API}/repos/{r}", soft=True)
            for repo, info in zip(top, pool.map(lookup, top)):
                if info:
                    stars[repo] = info.get("stargazers_count", 0)
        for r in rows:
            r["stars"] = stars.get(r["repo"])
            r["license"] = "?" if r["stars"] is None else r["license"]
        if min_stars:
            before = len(rows)
            # Unknown star counts survive the floor: dropping them would hide
            # projects purely because we ran out of anonymous API budget.
            rows = [r for r in rows if r["stars"] is None or r["stars"] >= min_stars]
            dropped["below star floor"] += before - len(rows)
        capped = len(by_hits) > ANON_STAR_LOOKUPS or RATE_LIMITED

    rows.sort(key=lambda r: (r["big"],
                            -(r["stars"] if r["stars"] is not None else -1),
                            -r["comments"]))
    return rows, capped, {k: v for k, v in dropped.items() if v}


def search_untagged(labels, per_phrase, big_owners, drop_big, per_repo, seen_urls):
    """Accessibility work nobody labeled.

    The new default label only helps once maintainers apply it. Everything filed
    before that is still sitting under `bug` or `enhancement`, so this searches
    issue text for the phrases that give it away and excludes anything already
    carrying an accessibility label.
    """
    excluded = ",".join(f'"{l}"' if " " in l else l for l in labels)
    rows, per_repo_count = [], {}

    for phrase in UNTAGGED_PHRASES:
        # in:title is the precision fix: matching the body pulled in issue
        # templates and game-compatibility reports that merely said the words.
        query = f'"{phrase}" in:title is:issue is:open is:public -label:{excluded}'
        for i in search_rest(query, per_phrase, sort="comments"):
            url = i["html_url"]
            if url in seen_urls:
                continue
            full = "/".join(i["repository_url"].split("/")[-2:])
            big = full.split("/")[0].lower() in big_owners
            if drop_big and big:
                continue
            if per_repo and per_repo_count.get(full, 0) >= per_repo:
                continue
            per_repo_count[full] = per_repo_count.get(full, 0) + 1
            seen_urls.add(url)
            issue_labels = [l["name"] for l in i.get("labels", [])]
            rows.append({
                "repo": full,
                "big": big,
                "area": classify(i["title"], issue_labels, i.get("body")),
                "source": f'text match: "{phrase}"',
                "stars": None,
                "license": "?",
                "language": "-",
                "number": i["number"],
                "title": i["title"],
                "url": url,
                "labels": issue_labels,
                "comments": i.get("comments", 0),
                "createdAt": i["created_at"],
                "updatedAt": i["updated_at"],
            })
    return rows


def add_stars(rows):
    """Fill in star counts for rows that arrived without them (GraphQL, batched)."""
    missing = sorted({r["repo"] for r in rows if r["stars"] is None})
    if not missing:
        return

    if TOKEN:
        # One GraphQL call per 50 repos beats one REST call each.
        for chunk_start in range(0, len(missing), 50):
            chunk = missing[chunk_start:chunk_start + 50]
            fields = []
            for idx, full in enumerate(chunk):
                owner, name = full.split("/", 1)
                fields.append(f'r{idx}: repository(owner: "{owner}", name: "{name}") '
                              '{ nameWithOwner stargazerCount licenseInfo { spdxId } '
                              'primaryLanguage { name } }')
            data = graphql("query { " + " ".join(fields) + " }", {}) or {}
            for value in data.values():
                if not value:
                    continue
                for r in rows:
                    if r["repo"] == value["nameWithOwner"]:
                        r["stars"] = value["stargazerCount"]
                        r["license"] = (value.get("licenseInfo") or {}).get("spdxId") or "none"
                        r["language"] = (value.get("primaryLanguage") or {}).get("name") or "-"
    else:
        for full in missing[:ANON_STAR_LOOKUPS]:
            info = request(f"{API}/repos/{full}", soft=True)
            if not info:
                break
            for r in rows:
                if r["repo"] == full:
                    r["stars"] = info.get("stargazers_count", 0)
                    r["language"] = info.get("language") or "-"


def refine_areas(rows, cap):
    """Read each issue and label it ourselves.

    The bulk search cannot carry issue bodies (GraphQL returns
    RESOURCE_LIMITS_EXCEEDED at 100 nodes and times out at 50), so this reads
    them one at a time: one REST call per issue, eight in flight, soft so a
    failure costs a row rather than the run. Reading the body is what turns
    "Improve accessibility" into "Keyboard & focus".
    """
    todo = rows[:cap]
    if not todo:
        return 0, 0

    def fetch(r):
        return request(f"{API}/repos/{r['repo']}/issues/{r['number']}", soft=True)

    moved, read = 0, 0
    with ThreadPoolExecutor(max_workers=8) as pool:
        for r, info in zip(todo, pool.map(fetch, todo)):
            if not info:
                continue
            read += 1
            before = r.get("area", GENERAL)
            areas = classify_all(r["title"], r["labels"], info.get("body"))
            r["areas"] = areas
            r["area"] = areas[0]
            if before == GENERAL and areas[0] != GENERAL:
                moved += 1
    return moved, read


def search_owner_issues(owners, labels):
    """Every accessibility issue in the owners' repos, open and closed."""
    merged = {}
    for owner in owners:
        query = f"{label_query(labels)} is:issue user:{owner}"
        for item in search_rest(query, 1000, sort="updated"):
            repo = "/".join(item["repository_url"].split("/")[-2:])
            merged[(repo, item["number"])] = {
                "repo": repo,
                "number": item["number"],
                "title": item["title"],
                "state": item["state"],
                "url": item["html_url"],
                "labels": [l["name"] for l in item.get("labels", [])],
                "assignee": ", ".join(a["login"] for a in item.get("assignees") or []) or "-",
                "comments": item.get("comments", 0),
                "createdAt": item["created_at"],
                "updatedAt": item["updated_at"],
            }
    return list(merged.values())


def fetch_owner_repos(owners):
    """Public repos always. Private ones too when the token owns them."""
    me = None
    if TOKEN:
        who = request(f"{API}/user")
        me = (who or {}).get("login")

    repos = []
    for owner in owners:
        if me and owner.lower() == me.lower():
            paths = ["/user/repos?affiliation=owner&per_page=100"]
        else:
            paths = [f"/users/{owner}/repos?per_page=100&type=owner",
                     f"/orgs/{owner}/repos?per_page=100&type=sources"]
        got = None
        for path in paths:
            page, collected = 1, []
            while True:
                data = request(f"{API}{path}&page={page}")
                if data is None:
                    break
                collected.extend(data)
                if len(data) < 100:
                    break
                page += 1
            if collected:
                got = collected
                break
        if not got:
            print(f"  warning: no repos found for '{owner}' (typo, or all private)",
                  file=sys.stderr)
            continue
        repos.extend(got)
    return repos


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------

def parse_ts(value):
    return datetime.fromisoformat(value.replace("Z", "+00:00")) if value else None


def build_workbook(global_rows, capped, dropped, mine, repos, owners, labels,
                   filter_note, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    open_fill = PatternFill("solid", fgColor="FFF2CC")    # amber: needs attention
    closed_fill = PatternFill("solid", fgColor="E2EFDA")  # green: done
    big_fill = PatternFill("solid", fgColor="EDEDED")     # grey: big company, hidden by default
    link_font = Font(color="0563C1", underline="single")

    def style(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def date_cells(ws, row, cols):
        """Real dates, not strings: Excel sorts text dates alphabetically."""
        for col in cols:
            ws.cell(row=row, column=col).number_format = "yyyy-mm-dd"

    def link_cell(ws, row, col, url):
        c = ws.cell(row=row, column=col)
        c.hyperlink = url
        c.value = "open"
        c.font = link_font

    now = datetime.now(timezone.utc)
    wb = Workbook()

    # --- Sheet 1: all public repos, most popular first ---------------------
    ws = wb.active
    ws.title = "All public repos"
    ws.append(["Area", "Also covers", "Owner", "Repo", "Stars", "License", "Language",
               "#", "Title", "Labels", "Found via", "Comments", "Good first issue",
               "Opened", "Last activity", "Age (days)", "Idle (days)", "Link"])
    big_rows = []
    for r in global_rows:
        created = parse_ts(r["createdAt"])
        updated = parse_ts(r["updatedAt"])
        gfi = any("good first issue" in l.lower() or "good-first-issue" in l.lower()
                  or "help wanted" in l.lower() for l in r["labels"])
        areas = r.get("areas") or [r.get("area", GENERAL)]
        ws.append([
            r.get("area", GENERAL),
            ", ".join(areas[1:]) or "-",
            "big company" if r.get("big") else "independent",
            r["repo"],
            r["stars"],
            r.get("license", "?"),
            r["language"],
            r["number"],
            r["title"],
            ", ".join(r["labels"]),
            r.get("source", "labeled"),
            r["comments"],
            "yes" if gfi else "no",
            created.date() if created else None,
            updated.date() if updated else None,
            (now - created).days if created else None,
            (now - updated).days if updated else None,
            r["url"],
        ])
        date_cells(ws, ws.max_row, (14, 15))
        link_cell(ws, ws.max_row, 18, r["url"])
        if r.get("big"):
            big_rows.append(ws.max_row)
            for c in range(1, 19):
                ws.cell(row=ws.max_row, column=c).fill = big_fill
    style(ws, [24, 26, 13, 28, 8, 11, 11, 7, 46, 22, 22, 10, 16, 12, 13, 11, 11, 8])

    # Column A opens filtered to "independent". Clearing the filter in Excel (or
    # unhiding the rows) brings the big-company issues back, which is the point:
    # the decision lives in the spreadsheet, not in how the file was generated.
    if big_rows:
        ws.auto_filter.add_filter_column(2, ["independent"], blank=False)
        for row_idx in big_rows:
            ws.row_dimensions[row_idx].hidden = True

    # --- Sheet 2: pick your specialty --------------------------------------
    # An index, not data: someone who does screen reader work should be able to
    # see their pile without scrolling the whole public sheet.
    ws_area = wb.create_sheet("By specialty")
    ws_area.append(["Area", "Primary area", "Also covers", "Total if you filter both",
                    "Projects", "Good first issue", "Unlabeled", "Most-starred project"])

    def blank_tally():
        return {"n": 0, "also": 0, "repos": {}, "gfi": 0, "untagged": 0}

    tally = {}
    for r in global_rows:
        if r.get("big"):
            continue
        areas = r.get("areas") or [r.get("area", GENERAL)]
        t = tally.setdefault(areas[0], blank_tally())
        t["n"] += 1
        t["repos"][r["repo"]] = max(t["repos"].get(r["repo"], 0), r["stars"] or 0)
        if any("good first issue" in l.lower() or "good-first-issue" in l.lower()
               or "help wanted" in l.lower() for l in r["labels"]):
            t["gfi"] += 1
        if not r.get("source", "labeled").startswith("labeled"):
            t["untagged"] += 1
        # Secondary areas counted separately, so the index matches what filtering
        # on "Also covers" actually returns.
        for secondary in areas[1:]:
            tally.setdefault(secondary, blank_tally())["also"] += 1

    for area, t in sorted(tally.items(), key=lambda kv: -(kv[1]["n"] + kv[1]["also"])):
        top = max(t["repos"], key=lambda k: t["repos"][k]) if t["repos"] else "-"
        ws_area.append([area, t["n"], t["also"], t["n"] + t["also"],
                        len(t["repos"]), t["gfi"], t["untagged"], top])
    style(ws_area, [28, 13, 12, 22, 10, 16, 11, 32])

    # --- Sheet 3: your own repos ------------------------------------------
    ws2 = wb.create_sheet("My repos")
    ws2.append(["Area", "Also covers", "Repo", "#", "Title", "State", "Labels", "Assignee",
                "Comments", "Opened", "Last activity", "Age (days)", "Idle (days)", "Link"])
    # Open first, then most recently updated: the top of the sheet is the work.
    mine.sort(key=lambda i: (i["state"] != "open", -parse_ts(i["updatedAt"]).timestamp()))
    for i in mine:
        created = parse_ts(i["createdAt"])
        updated = parse_ts(i["updatedAt"])
        areas = i.get("areas") or [classify(i["title"], i["labels"])]
        ws2.append([
            areas[0], ", ".join(areas[1:]) or "-",
            i["repo"], i["number"], i["title"], i["state"],
            ", ".join(i["labels"]), i["assignee"], i["comments"],
            created.date() if created else None,
            updated.date() if updated else None,
            (now - created).days if created else None,
            (now - updated).days if updated else None,
            i["url"],
        ])
        row = ws2.max_row
        fill = open_fill if i["state"] == "open" else closed_fill
        for c in range(1, 15):
            ws2.cell(row=row, column=c).fill = fill
        date_cells(ws2, row, (10, 11))
        link_cell(ws2, row, 14, i["url"])
    style(ws2, [24, 24, 26, 6, 50, 8, 28, 16, 10, 12, 13, 11, 11, 8])

    # --- Sheet 4: your repo rollup ----------------------------------------
    ws3 = wb.create_sheet("My repo rollup")
    ws3.append(["Repo", "Stars", "Open a11y", "Closed a11y", "Total",
                "Private", "Archived", "Repo updated", "Link"])
    counts = {}
    for i in mine:
        b = counts.setdefault(i["repo"], {"open": 0, "closed": 0})
        b["open" if i["state"] == "open" else "closed"] += 1
    for repo in sorted(repos, key=lambda r: (
            -counts.get(r["full_name"], {"open": 0})["open"],
            -r.get("stargazers_count", 0),
            r["full_name"].lower())):
        c = counts.get(repo["full_name"], {"open": 0, "closed": 0})
        updated = parse_ts(repo.get("updated_at"))
        ws3.append([
            repo["full_name"], repo.get("stargazers_count", 0),
            c["open"], c["closed"], c["open"] + c["closed"],
            "yes" if repo.get("private") else "no",
            "yes" if repo.get("archived") else "no",
            updated.date() if updated else None,
            repo["html_url"],
        ])
        date_cells(ws3, ws3.max_row, (8,))
        if c["open"]:
            for col in range(1, 10):
                ws3.cell(row=ws3.max_row, column=col).fill = open_fill
        link_cell(ws3, ws3.max_row, 9, repo["html_url"])
    style(ws3, [34, 8, 11, 12, 8, 9, 10, 14, 8])

    # --- Sheet 5: run metadata --------------------------------------------
    ws4 = wb.create_sheet("Run info")
    open_mine = sum(1 for i in mine if i["state"] == "open")
    notes = [
        ("Last run (UTC)", now.strftime("%Y-%m-%d %H:%M")),
        ("Owners scanned", ", ".join(owners)),
        ("Labels matched", ", ".join(labels)),
        ("Signed in", "yes (private repos included)" if TOKEN else "no (public data only)"),
        ("Public issues listed", f"{sum(1 for r in global_rows if not r.get('big'))} independent"
                                 + (f", {sum(1 for r in global_rows if r.get('big'))} big company "
                                    f"(hidden by the column A filter)"
                                    if any(r.get('big') for r in global_rows) else "")),
        ("Public sample", "most-commented open accessibility issues, then sorted by stars"),
        ("Public filter", filter_note),
        ("Popularity metric", "repo stars (GitHub does not expose download counts)"),
        ("Your repos scanned", len(repos)),
        ("Your issues found", f"{len(mine)} ({open_mine} open)"),
    ]
    if dropped:
        notes.append(("Excluded from public sheet",
                      ", ".join(f"{v} {k}" for k, v in dropped.items())))
    if capped:
        notes.append(("Note", "some star counts are missing (anonymous rate limit); "
                              "sign in for a complete ranking"))
    for k, v in notes:
        ws4.append([k, v])
    for cell in ws4["A"]:
        cell.font = Font(bold=True)
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 70

    wb.save(out_path)
    return open_mine


# --------------------------------------------------------------------------

def main():
    global TOKEN

    ap = argparse.ArgumentParser(
        description="Track accessibility-labeled GitHub issues in a spreadsheet.")
    ap.add_argument("--owner", default=None,
                    help="GitHub user or org (comma-separated for several)")
    ap.add_argument("--out", default=DEFAULT_OUT, help="output .xlsx path")
    ap.add_argument("--labels", default=",".join(DEFAULT_LABELS),
                    help="comma-separated label spellings to match")
    ap.add_argument("--global-limit", type=int, default=300,
                    help="how many public issues to pull (max 1000, default 300)")
    ap.add_argument("--no-global", action="store_true",
                    help="skip the all-public-repos sheet")
    ap.add_argument("--no-deep-classify", action="store_true",
                    help="skip reading issue bodies (faster, but vaguer labels)")
    ap.add_argument("--classify-cap", type=int, default=600,
                    help="most issues to read in full for labeling, default 600")
    ap.add_argument("--no-untagged", action="store_true",
                    help="skip the search for accessibility work nobody labeled")
    ap.add_argument("--untagged-per-phrase", type=int, default=25,
                    help="how many results per untagged search phrase, default 25")
    ap.add_argument("--exclude-big-tech", action="store_true",
                    help="leave big-company issues out of the file entirely, rather than "
                         "including them behind the spreadsheet filter")
    ap.add_argument("--exclude-owners", default="",
                    help="extra owners to leave out of the public sheet")
    ap.add_argument("--any-license", action="store_true",
                    help="include public repos with no recognizable open source license")
    ap.add_argument("--min-stars", type=int, default=0,
                    help="ignore public repos below this star count")
    ap.add_argument("--per-repo", type=int, default=3,
                    help="max issues shown per repo on the public sheet, 0 for no cap")
    args = ap.parse_args()

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        sys.exit("Missing dependency. Install it with:\n  python3 -m pip install --user openpyxl")

    TOKEN = resolve_token()

    owner_arg = args.owner or os.environ.get("A11Y_OWNER")
    if not owner_arg and TOKEN:
        who = request(f"{API}/user")
        owner_arg = (who or {}).get("login")
    if not owner_arg:
        sys.exit("Which GitHub account should I scan? Pass --owner <username>.")
    owners = [o.strip() for o in owner_arg.split(",") if o.strip()]
    labels = [l.strip() for l in args.labels.split(",") if l.strip()]
    out_path = os.path.expanduser(args.out)

    print(f"Scanning {', '.join(owners)}" + ("" if TOKEN else " (not signed in, public data only)"))

    big_owners = set(BIG_TECH_OWNERS)
    big_owners |= {o.strip().lower() for o in args.exclude_owners.split(",") if o.strip()}

    parts = []
    if args.exclude_big_tech:
        parts.append(f"{len(big_owners)} big-company owners left out of the file")
    else:
        parts.append(f"{len(big_owners)} big-company owners included but hidden "
                     f"behind the column A filter")
    if not args.any_license:
        parts.append("open source license required where known")
    parts.append("no archived repos or forks")
    if args.per_repo:
        parts.append(f"at most {args.per_repo} issues per repo")
    if args.min_stars:
        parts.append(f"at least {args.min_stars} stars")
    filter_note = "; ".join(parts) if parts else "none"

    global_rows, capped, dropped = ([], False, {})
    if not args.no_global:
        global_rows, capped, dropped = search_global(
            labels, min(args.global_limit, 1000), big_owners, args.exclude_big_tech,
            not args.any_license, args.min_stars, args.per_repo)
        indie = sum(1 for r in global_rows if not r["big"])
        extra = len(global_rows) - indie
        print(f"  {indie} open accessibility issues in independent public repos"
              + (f", plus {extra} big-company ones hidden behind the filter" if extra else ""))

        if not args.no_untagged:
            seen = {r["url"] for r in global_rows}
            untagged = search_untagged(labels, args.untagged_per_phrase, big_owners,
                                       args.exclude_big_tech, args.per_repo, seen)
            add_stars(untagged)
            global_rows.extend(untagged)
            print(f"  {len(untagged)} more that look like accessibility work "
                  f"but carry no accessibility label")

        if not args.no_deep_classify:
            moved, read = refine_areas(global_rows, args.classify_cap)
            print(f"  read {read} issues to label them, {moved} of which no title could place")

        global_rows.sort(key=lambda r: (r["big"],
                                        -(r["stars"] if r["stars"] is not None else -1),
                                        -r["comments"]))

    mine = search_owner_issues(owners, labels)
    repos = fetch_owner_repos(owners)
    open_mine = build_workbook(global_rows, capped, dropped, mine, repos, owners,
                               labels, filter_note, out_path)

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"[{stamp}] your repos: {len(mine)} items ({open_mine} open) across "
          f"{len(repos)} repos -> {out_path}")


if __name__ == "__main__":
    main()
